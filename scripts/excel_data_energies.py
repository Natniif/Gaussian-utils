#This is a script that takes a gaussian output optimisation file and returns the mulliken charges and 
#spin densities for each atom as well as the charge and spin densities for the specific sections in the 
#substrate

'''
This script takes 3 inputs: 
    [1]: gaussian output file
    [2]: Name of excel file you want to add data to
    [3]: Name of tab you want to add to the excel workbook

    call code like this: python3 excel_data_energies.py [1] [2] [3] 
'''

from openpyxl import Workbook 
from openpyxl import load_workbook
import sys

gauss_out = open(sys.argv[1],'r')       #opens the Gaussian output file for reading

workbook = load_workbook(sys.argv[2])
workbook.create_sheet(sys.argv[3])

sheet = workbook[sys.argv[3]] 

listalines = gauss_out.readlines()

last = max(idx for idx, val in enumerate(listalines) if val.strip() == 'Mulliken charges and spin densities:')

vals = []
for i in range(last+2, last + 110):
    vals.append(listalines[i].split())


for i in range(len(vals)):
    atom_num_cell = 'C' + str(int(4 + i))
    atom_label_cell = 'D' + str(int(4 + i))
    atom_spin_cell = 'E' + str(int(4 + i))
    atom_charge_cell = 'F' + str(int(4 + i))

    sheet[atom_num_cell] = vals[i][0] 
    sheet[atom_label_cell] = vals[i][1]
    sheet[atom_spin_cell].value = float(vals[i][2])
    sheet[atom_charge_cell].value = float(vals[i][3])


sheet['e3'] = 'Mulliken Charges'
sheet['F3'] = 'Spin Densities'
sheet["I3"] = "Axial ligand"
sheet['j3'] = 'Fe'
sheet['k3'] = 'Oxygen'
sheet['l3'] = 'Substrate'
sheet['m3'] = 'Protein'
sheet['n3'] = 'Sum'
sheet['h4'] = 'Spin Density' 
sheet['h5'] = 'Charge'
sheet['i4'] = '= SUM(F35:F45) + F109'
sheet['j4'] = '=F46'
sheet['k4'] = '=F47'
sheet['l4'] = '=SUM(F54:F106)'
sheet['m4'] = '=SUM(F3:F34)+SUM(F48:F53)+F107+F108+F110+F111'
sheet['n4'] = '=SUM(I4:M4)'
sheet['i5'] = '= SUM(E35:E45) + E109'
sheet['j5'] = '=E46'
sheet['k5'] = '=E47'
sheet['l5'] = '=SUM(E54:E106)'
sheet['m5'] = '=SUM(E3:E34)+SUM(E48:E53)+E107+E108+E110+E111'
sheet['n5'] = '=SUM(I5:M5)'

workbook.save(sys.argv[2])
