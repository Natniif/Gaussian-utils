#### EXTRACT.PY - This program takes as input an output file from a geometry optimization/frequency calculation run in Gaussian and
####             extracts a) its energy; b) the energy's zero-point correction; c) the sum of free energies; d) charges and spin densities.
#### Usage: in a terminal, type in: python3 extract.py GAUSSIAN_OUTPUT_FILE NAME_OF_THE_FILE_WHERE_DATA_WILL_BE_EXTRACTED_TO and hit ENTER

import sys
gauss_out = open(sys.argv[1],'r')       #opens the Gaussian output file for reading
#data = open(sys.argv[2],'w')       #creates a file for writing all the important information

from openpyxl import Workbook
from openpyxl.utils import FORMULAE

filename = sys.argv[2]
workbook = Workbook()
sheet = workbook.active

wb = load_workbook(sys.argv[2])
wb.create_sheet(sys.argv[3])


listalines = gauss_out.readlines()
last = max(idx for idx, val in enumerate(listalines) if val.strip() == 'Mulliken charges and spin densities:')

vals = []
for i in range(last+1,last+110):              #IMPORTANT: instead of '224', take the number of atoms of your system plus 2
    vals.append(listalines[i].split())

for i in range(len(vals)):
    val = i + 4
    sheet['C',val] = vals[i][0]
    sheet['D',val] = vals[i][1]
    sheet['E',val] = vals[i][2]
    sheet['F',val] = vals[i][3]


sheet['I3'] = "Axial Ligand"
sheet['j3'] = 'Fe'
sheet['k3'] = 'Oxygen'
sheet['l3'] = 'Substrate'
sheet['m3'] = 'Protein'
sheet['n3'] = 'Sum'
sheet['h5'] = 'Charge'
sheet['h6'] = 'Spin Density'
sheet['e3'] = 'Mulliken Charges'
sheet['f3'] = 'Spin Densities'

sheet['i4'] = '= SUM(F35:F45) + F109'
sheet['j4'] = '=F46'
sheet['k4'] = '=F47'
sheet['l4'] = '=SUM(F54:F106)'
sheet['m4'] = '=SUM(F3:F34)+SUM(F48:F53)+F107+F108+F110+F111'
sheet['n4'] = '=SUM(I4:M4)'

sheet['i5'] = '= SUM(e35:e45) + e109'
sheet['j5'] = '=e46'
sheet['k5'] = '=e47'
sheet['l5'] = '=SUM(e54:e106)'
sheet['m5'] = '=SUM(e3:e34)+SUM(e48:e53)+e107+e108+e110+e111'
sheet['n5'] = '=SUM(I5:M5)'


gauss_out.close()
workbook.save(filename=filename)


