'''input requires: 
    [1] output file
    [2] scan step 
    [3] starting scan distance
    [4] microsoft excel sheet you want data added to 
    [5] name you want to give sheet in excel'''


'''TO DO
Add code in that exports data straight to excel sheet

Add in code that find the distance between two points specified - maybe another script
'''

import sys
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

# This class is redundant i think
# class ScanPoint:
#     def __init__(self, scan_num):
#         self.scan_num = scan_num

# More things that can be done?
class Scan:
    def __init__(self, energies, scan_distances):
        self.energies = energies
        self.scan_distances = scan_distances

    def max_scan(self):
        self.max_point = max(self.energies) 
        index_of_max = self.energies.index(self.max_point)
        return self.max_point , self.scan_distances[index_of_max] 

    def plot(self):
        plt.figure(1)

        plt.subplot(121)
        plt.plot(self.scan_distances,self.energies)
        plt.xlabel("Scan Distance")
        plt.ylabel("Energy")
        plt.title("Absolute Energy vs Scan Distance for " + sys.argv[1])

        plt.subplot(122)
        plt.plot(self.scan_distances,self.relative_energy)
        plt.xlabel("Scan Distance")
        plt.ylabel("Relative Energy")
        plt.title("Relative Energy vs Scan Distance for " + sys.argv[1])

        plt.show()
    def relative_energies(self):
        self.relative_energy = []
        for i in range(len(self.energies)):
            self.relative_energy.append(float(self.energies[i]) - float(self.max_point))
        return(self.relative_energy)


def excel_output(scan_num,scan_steps,energies,relative_energy, excel_sheet,sheet_name):
    workbook = load_workbook(excel_sheet) 
    workbook.create_sheet(sheet_name) 

    sheet = workbook[sheet_name]
    
    sheet['B2'] = 'Scan Step'
    sheet['C2'] = 'R'
    sheet['D2'] = 'Energy A.U'
    sheet['E2'] = 'DE'

    for i in range(len(scan_steps)):
        column_b_name = 'B' + str(int(3 + i))
        column_c_name = 'C' + str(int(3 + i))
        column_d_name = 'D' + str(int(3 + i))
        column_e_name = 'E' + str(int(3 + i))
        
        sheet[column_b_name].value = scan_num[i]
        sheet[column_c_name].value = scan_steps[i] 
        sheet[column_d_name].value = energies[i]
        sheet[column_e_name].value = relative_energy[i]

    workbook.save(excel_sheet)



# This is really messy...
# Could add more of this into the class section but whatever it works fine quickly anyways
def extractor(file):
    energies = []
    scan_num = []
    with open(file, 'r') as f:
        j = 0
        lines = f.readlines()
        for line in lines:
            counter = 0
            if 'Converged?' in line:
                # print(line, j)
                if len(lines[j+1]) > 4 and len(lines[j+2]) > 4 and len(lines[j+3]) > 4 and len(lines[j+4]) > 4:
                    if 'YES' in lines[j+1] and 'YES' in lines[j+2] and 'YES' in lines[j+3] and 'YES' in lines[j+4]:
                        print('found converged scan on line:', j+1)
                        for i in range(j+1, len(lines)):
                            line_lst = lines[i].split()
                                                    
                            # Search for initialization step
                            if 'Initialization' in line_lst:
                                # print('startring lines found')
                                break
                           
                            # Search for scan step
                            if 'scan point ' in lines[i]:
                                pass
                                line_lst1 = lines[i].split()
                                scan_num.append(line_lst1[-4])
                                print(line_lst1[-4])
                            
                            # Search for energy within scan step
                            if 'SCF Done:' in lines[i]:
                                line_lst2 = lines[i].split()
                                energies.append(line_lst2[-5])
                                print(line_lst2[-5])

                            # End when new scan step
                            if 'Converged?' in lines[i]:
                                # print('Took ' + str(counter) + ' scan steps')
                                break       
                    else: 
                        counter += 1
                        pass
                j += 1         
            else:
                j +=1
                    
    # print(energies)
    # print(scan_num)
    f.close()

    scan_distance_inp = float(sys.argv[2])
    scan_start_distance = float(sys.argv[3])
    scan_distance = [scan_start_distance]
    
    for i in range(1,len(scan_num)):
        scan_distance.append(round(scan_start_distance - i*scan_distance_inp, 4))

    # print(scan_distance)

    out = Scan(energies, scan_distance)
    print(out.max_scan())
    x = out.relative_energies()
    out.plot()
    excel_output(scan_num,scan_distance,energies,x,sys.argv[4],sys.argv[5])

def last_scan_point(file):
    with open(file, 'r') as f:
        lines = f.readlines()
        last_coordinate = 0
        j = 1
        for line in lines:
            if 'Coordinates' in line:
                if j > last_coordinate:
                    last_coordinate = j
                j += 1
            else: 
                j += 1
        coordinates = '' 
        for i in range(last_coordinate+2, len(lines)):
            if '------------' in lines[i]:
                break
            else:
                coord_list = lines[i].split()
                coordinates += ' '
                coordinates += str(coord_list[1]) + '  '
                coordinates += str(coord_list[-3]) + '   '
                coordinates += str(coord_list[-2]) + '   '
                coordinates += str(coord_list[-1]) + '   '
                coordinates += '\n'

        return coordinates




if __name__ == "__main__":
    print(extractor(sys.argv[1]))
    # print(last_scan_point(sys.argv[1]))
    
